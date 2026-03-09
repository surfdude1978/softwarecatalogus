"""Tests voor export functionaliteit (CSV, Excel, AMEFF)."""
import csv
import io

import pytest
from django.urls import reverse

pytestmark = pytest.mark.django_db


class TestExportPakkettenCSV:
    def test_csv_export_publiek_beschikbaar(self, api_client, pakket):
        url = reverse("api:export-pakketten-csv")
        response = api_client.get(url)
        assert response.status_code == 200
        assert "text/csv" in response["Content-Type"]
        assert "pakketten_" in response["Content-Disposition"]

    def test_csv_bevat_pakket_data(self, api_client, pakket):
        url = reverse("api:export-pakketten-csv")
        response = api_client.get(url)
        content = response.content.decode("utf-8-sig")
        assert "Suite4Gemeenten" in content
        assert "Centric" in content

    def test_csv_header_rij(self, api_client, pakket):
        url = reverse("api:export-pakketten-csv")
        response = api_client.get(url)
        content = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content), delimiter=";")
        header = next(reader)
        assert "Naam" in header
        assert "Leverancier" in header
        assert "Licentievorm" in header


class TestExportPakkettenExcel:
    def test_excel_export(self, api_client, pakket):
        url = reverse("api:export-pakketten-xlsx")
        response = api_client.get(url)
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_excel_bevat_data(self, api_client, pakket):
        import openpyxl
        url = reverse("api:export-pakketten-xlsx")
        response = api_client.get(url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        # Check header row
        assert ws.cell(1, 1).value == "Naam"
        # Check data row
        assert ws.cell(2, 1).value == "Suite4Gemeenten"


class TestExportPakketOverzichtCSV:
    def test_overzicht_csv_vereist_auth(self, api_client):
        url = reverse("api:export-overzicht-csv")
        response = api_client.get(url)
        assert response.status_code == 401

    def test_overzicht_csv_met_auth(self, auth_client, pakket_gebruik):
        url = reverse("api:export-overzicht-csv")
        response = auth_client.get(url)
        assert response.status_code == 200
        content = response.content.decode("utf-8-sig")
        assert "Suite4Gemeenten" in content

    def test_csv_bevat_bom(self, auth_client, pakket_gebruik):
        url = reverse("api:export-overzicht-csv")
        response = auth_client.get(url)
        assert response.content.startswith(b"\xef\xbb\xbf")

    def test_logt_export_actie(self, auth_client, pakket_gebruik):
        from apps.core.audit import AuditLog
        url = reverse("api:export-overzicht-csv")
        auth_client.get(url)
        assert AuditLog.objects.filter(
            actie=AuditLog.Actie.EXPORT,
            object_type="PakketOverzicht",
        ).exists()

    def test_fout_zonder_organisatie(self, api_client, db):
        from rest_framework_simplejwt.tokens import RefreshToken

        from apps.gebruikers.models import User
        user = User.objects.create_user(
            email="noorg_csv@test.nl",
            password="test",
            naam="No Org",
            organisatie=None,
            rol="gebruik_beheerder",
        )
        token = str(RefreshToken.for_user(user).access_token)
        api_client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        url = reverse("api:export-overzicht-csv")
        response = api_client.get(url)
        assert response.status_code == 400


class TestExportPakketOverzichtExcel:
    def test_excel_export_vereist_auth(self, api_client):
        url = reverse("api:export-overzicht-xlsx")
        response = api_client.get(url)
        assert response.status_code == 401

    def test_excel_export_met_auth(self, auth_client, pakket_gebruik):
        url = reverse("api:export-overzicht-xlsx")
        response = auth_client.get(url)
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_excel_bevat_pakket_data(self, auth_client, pakket_gebruik):
        import openpyxl
        url = reverse("api:export-overzicht-xlsx")
        response = auth_client.get(url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Pakketoverzicht"]
        waarden = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert pakket_gebruik.pakket.naam in waarden

    def test_content_disposition_header(self, auth_client, pakket_gebruik):
        url = reverse("api:export-overzicht-xlsx")
        response = auth_client.get(url)
        assert "attachment" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

    def test_fout_zonder_organisatie(self, api_client, db):
        from rest_framework_simplejwt.tokens import RefreshToken

        from apps.gebruikers.models import User
        user = User.objects.create_user(
            email="noorg_excel@test.nl",
            password="test",
            naam="No Org",
            organisatie=None,
            rol="gebruik_beheerder",
        )
        token = str(RefreshToken.for_user(user).access_token)
        api_client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        url = reverse("api:export-overzicht-xlsx")
        response = api_client.get(url)
        assert response.status_code == 400


class TestExportPakketOverzichtAMEFF:
    def test_ameff_export_vereist_auth(self, api_client):
        url = reverse("api:export-overzicht-ameff")
        response = api_client.get(url)
        assert response.status_code == 401

    def test_ameff_export_retourneert_xml(self, auth_client, pakket_gebruik, gemma_component):
        from apps.architectuur.models import PakketGemmaComponent
        PakketGemmaComponent.objects.create(pakket=pakket_gebruik.pakket, gemma_component=gemma_component)
        url = reverse("api:export-overzicht-ameff")
        response = auth_client.get(url)
        assert response.status_code == 200
        assert "application/xml" in response["Content-Type"]
        content = response.content.decode("utf-8")
        assert "archimate" in content.lower() or "model" in content.lower()
