"""Export views voor CSV, Excel en AMEFF."""
import csv
import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView

from apps.api.permissions import IsFullyAuthenticated, IsFunctioneelBeheerder
from apps.architectuur.ameff_export import generate_ameff
from apps.core.audit import AuditLog, log_actie
from apps.pakketten.models import Pakket, PakketGebruik


class ExportPakkettenCSV(APIView):
    """Exporteer pakketten als CSV."""
    permission_classes = [AllowAny]

    def get(self, request):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="pakketten_{datetime.now():%Y%m%d}.csv"'
        response.write("\ufeff")  # BOM voor Excel compatibiliteit

        writer = csv.writer(response, delimiter=";")
        writer.writerow([
            "Naam", "Versie", "Status", "Leverancier", "Licentievorm",
            "Website", "Beschrijving", "Aantal gemeenten",
        ])

        pakketten = Pakket.objects.filter(
            status__in=["actief", "concept"]
        ).select_related("leverancier")

        for p in pakketten:
            writer.writerow([
                p.naam,
                p.versie,
                p.get_status_display(),
                p.leverancier.naam if p.leverancier else "",
                p.get_licentievorm_display(),
                p.website_url,
                p.beschrijving[:500],
                p.pakketgebruik_set.filter(status="in_gebruik").count(),
            ])

        return response


class ExportPakkettenExcel(APIView):
    """Exporteer pakketten als Excel (.xlsx)."""
    permission_classes = [AllowAny]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pakketten"

        # Header
        headers = [
            "Naam", "Versie", "Status", "Leverancier", "Licentievorm",
            "Website", "Beschrijving", "Aantal gemeenten",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        pakketten = Pakket.objects.filter(
            status__in=["actief", "concept"]
        ).select_related("leverancier")

        for p in pakketten:
            ws.append([
                p.naam,
                p.versie,
                p.get_status_display(),
                p.leverancier.naam if p.leverancier else "",
                p.get_licentievorm_display(),
                p.website_url,
                p.beschrijving[:500],
                p.pakketgebruik_set.filter(status="in_gebruik").count(),
            ])

        # Auto-breedte kolommen
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="pakketten_{datetime.now():%Y%m%d}.xlsx"'
        return response


class ExportPakketOverzichtCSV(APIView):
    """Exporteer eigen pakketoverzicht als CSV."""
    permission_classes = [IsFullyAuthenticated]

    def get(self, request):
        user = request.user
        if not user.organisatie:
            return HttpResponse("Geen organisatie gekoppeld.", status=400)

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="pakketoverzicht_{user.organisatie.naam}_{datetime.now():%Y%m%d}.csv"'
        )
        response.write("\ufeff")

        writer = csv.writer(response, delimiter=";")
        writer.writerow(["Pakket", "Versie", "Status gebruik", "Leverancier", "Startdatum", "Notitie"])

        gebruik = PakketGebruik.objects.filter(
            organisatie=user.organisatie
        ).select_related("pakket", "pakket__leverancier")

        for pg in gebruik:
            writer.writerow([
                pg.pakket.naam,
                pg.pakket.versie,
                pg.get_status_display(),
                pg.pakket.leverancier.naam if pg.pakket.leverancier else "",
                pg.start_datum.isoformat() if pg.start_datum else "",
                pg.notitie,
            ])

        log_actie(
            request,
            AuditLog.Actie.EXPORT,
            object_type="PakketOverzicht",
            object_id=str(user.organisatie.pk),
            object_omschrijving=user.organisatie.naam,
            extra={"formaat": "csv", "records": gebruik.count()},
        )

        return response


class ExportPakketOverzichtExcel(APIView):
    """Exporteer eigen pakketoverzicht als Excel (.xlsx)."""
    permission_classes = [IsFullyAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        user = request.user
        if not user.organisatie:
            return HttpResponse("Geen organisatie gekoppeld.", status=400)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pakketoverzicht"

        # Titelrij
        org_naam = user.organisatie.naam
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Pakketoverzicht — {org_naam}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="left")

        ws["A2"] = f"Geëxporteerd op: {datetime.now():%d-%m-%Y %H:%M}"
        ws["A2"].font = Font(italic=True, size=10, color="888888")
        ws.append([])  # Lege rij

        # Header
        headers = [
            "Pakket", "Versie", "Leverancier", "Status gebruik",
            "Startdatum", "Einddatum", "Notitie",
            "GEMMA-componenten", "Standaarden",
        ]
        header_row = ws.max_row + 1
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F5C99")
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        gebruik = PakketGebruik.objects.filter(
            organisatie=user.organisatie
        ).select_related(
            "pakket", "pakket__leverancier"
        ).prefetch_related(
            "pakket__gemma_componenten",
            "pakket__standaarden",
        ).order_by("pakket__naam")

        for pg in gebruik:
            gemma_namen = ", ".join(
                gc.naam for gc in pg.pakket.gemma_componenten.all()
            )
            standaard_namen = ", ".join(
                s.naam for s in pg.pakket.standaarden.all()
            )
            ws.append([
                pg.pakket.naam,
                pg.pakket.versie or "",
                pg.pakket.leverancier.naam if pg.pakket.leverancier else "",
                pg.get_status_display(),
                pg.start_datum.strftime("%d-%m-%Y") if pg.start_datum else "",
                pg.eind_datum.strftime("%d-%m-%Y") if pg.eind_datum else "",
                pg.notitie or "",
                gemma_namen,
                standaard_namen,
            ])

        # Kolombreedtes aanpassen
        col_widths = [35, 12, 30, 15, 12, 12, 40, 40, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Bevroren headerrij
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        org_slug = org_naam.lower().replace(" ", "_")[:30]
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="pakketoverzicht_{org_slug}_{datetime.now():%Y%m%d}.xlsx"'
        )
        return response


class ExportPakketOverzichtAMEFF(APIView):
    """Exporteer eigen pakketoverzicht als AMEFF (ArchiMate Exchange)."""
    permission_classes = [IsFullyAuthenticated]

    def get(self, request):
        user = request.user
        if not user.organisatie:
            return HttpResponse("Geen organisatie gekoppeld.", status=400)

        xml_content = generate_ameff(organisatie_id=user.organisatie.id)

        response = HttpResponse(xml_content, content_type="application/xml")
        response["Content-Disposition"] = (
            f'attachment; filename="pakketoverzicht_{user.organisatie.naam}_{datetime.now():%Y%m%d}.ameff"'
        )
        return response


class ExportAuditLogCSV(APIView):
    """
    Exporteer de audit log als CSV.

    Uitsluitend toegankelijk voor de **functioneel beheerder**.
    Ondersteunt optionele filters via query parameters:

    * ``actie`` — filter op actie (bijv. ``aangemaakt``, ``verwijderd``)
    * ``actor_email`` — filter op e-mailadres van de actor
    * ``object_type`` — filter op objecttype (bijv. ``Pakket``)
    * ``van`` — begin datum/tijd (ISO, bijv. ``2025-01-01``)
    * ``tot`` — eind datum/tijd (ISO, bijv. ``2025-12-31``)
    """

    permission_classes = [IsFunctioneelBeheerder]

    def get(self, request):
        from datetime import datetime as _dt
        from zoneinfo import ZoneInfo

        from django.utils.dateparse import parse_date, parse_datetime
        from django.utils.timezone import make_aware

        ams = ZoneInfo("Europe/Amsterdam")
        qs = AuditLog.objects.all()

        # ── Filters ─────────────────────────────────────────────────────────
        actie = request.query_params.get("actie")
        if actie:
            qs = qs.filter(actie=actie)

        actor_email = request.query_params.get("actor_email")
        if actor_email:
            qs = qs.filter(actor_email__icontains=actor_email)

        object_type = request.query_params.get("object_type")
        if object_type:
            qs = qs.filter(object_type__iexact=object_type)

        van = request.query_params.get("van")
        if van:
            dt = parse_datetime(van)
            if not dt:
                d = parse_date(van)
                if d:
                    dt = make_aware(_dt(d.year, d.month, d.day, 0, 0, 0), ams)
            if dt:
                qs = qs.filter(tijdstip__gte=dt)

        tot = request.query_params.get("tot")
        if tot:
            dt = parse_datetime(tot)
            if not dt:
                d = parse_date(tot)
                if d:
                    dt = make_aware(_dt(d.year, d.month, d.day, 23, 59, 59), ams)
            if dt:
                qs = qs.filter(tijdstip__lte=dt)

        # ── Bouw CSV ─────────────────────────────────────────────────────────
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="auditlog_{datetime.now():%Y%m%d_%H%M}.csv"'
        )
        response.write("\ufeff")  # BOM voor Excel-compatibiliteit

        writer = csv.writer(response, delimiter=";")
        writer.writerow([
            "Tijdstip", "Actor e-mail", "Actor ID",
            "Actie", "Object type", "Object ID", "Object",
            "IP-adres", "User-agent",
        ])

        for entry in qs.order_by("-tijdstip"):
            writer.writerow([
                entry.tijdstip.strftime("%Y-%m-%d %H:%M:%S"),
                entry.actor_email,
                entry.actor_id or "",
                entry.actie,
                entry.object_type,
                entry.object_id or "",
                entry.object_omschrijving,
                entry.ip_adres or "",
                entry.user_agent[:100],
            ])

        # Log de export zelf ook
        log_actie(
            request,
            AuditLog.Actie.EXPORT,
            object_type="AuditLog",
            object_omschrijving=f"CSV export ({qs.count()} regels)",
            extra={
                "formaat": "csv",
                "records": qs.count(),
                "filters": dict(request.query_params),
            },
        )

        return response
